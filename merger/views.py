import json
from django.shortcuts import render
from django.http import HttpResponse
from PyPDF2 import PdfMerger, PdfReader
import csv
import io
import openpyxl
import os
from .utils import docx_to_text, text_to_pdf_buffer

def index(request):
    if request.method == 'POST':
        job_name = request.POST.get('job_name', 'Merged_Document')
        if not job_name.strip(): job_name = "Merged_Document"
        
        merge_type = request.POST.get('merge_type')
        files = request.FILES.getlist('files')
        file_order = request.POST.get('file_order', '')
        page_ranges_json = request.POST.get('page_ranges', '[]')

        if not files:
            return render(request, 'merger/index.html', {'error': 'No files uploaded.'})

        # Reorder files based on file_order
        try:
            if file_order:
                order_indices = [int(i) for i in file_order.split(',')]
                valid_indices = [i for i in order_indices if 0 <= i < len(files)]
                ordered_files = [files[i] for i in valid_indices]
                page_ranges = json.loads(page_ranges_json)
            else:
                ordered_files = files
                page_ranges = ['all'] * len(ordered_files)
        except:
            ordered_files = files
            page_ranges = ['all'] * len(ordered_files)

        # Perform merge
        try:
            if merge_type == 'PDF':
                merged_content = merge_pdfs(ordered_files, page_ranges)
                content_type = 'application/pdf'
                filename = f"{job_name}.pdf"
            elif merge_type == 'CSV':
                merged_content = merge_csvs(ordered_files)
                content_type = 'text/csv'
                filename = f"{job_name}.csv"
            elif merge_type == 'EXCEL':
                merged_content = merge_excels(ordered_files)
                content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                filename = f"{job_name}.xlsx"
            else:
                return render(request, 'merger/index.html', {'error': 'Unsupported merge type.'})

            response = HttpResponse(merged_content, content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return render(request, 'merger/index.html', {'error': f'Merge failed: {str(e)}'})

    return render(request, 'merger/index.html')

def parse_page_range(range_str, max_pages):
    if not range_str or range_str.lower() == 'all':
        return range(max_pages)
    pages = []
    for part in range_str.split(','):
        part = part.strip()
        if '-' in part:
            try:
                start, end = map(int, part.split('-'))
                for p in range(start - 1, end):
                    if 0 <= p < max_pages: pages.append(p)
            except: continue
        else:
            try:
                p = int(part) - 1
                if 0 <= p < max_pages: pages.append(p)
            except: continue
    return pages if pages else range(max_pages)

def merge_pdfs(files, page_ranges):
    merger = PdfMerger()
    
    try:
        for i, f in enumerate(files):
            ext = f.name.split('.')[-1].lower()
            range_str = page_ranges[i] if i < len(page_ranges) else 'all'

            if ext == 'pdf':
                reader = PdfReader(f)
                max_pages = len(reader.pages)
                pages_to_include = parse_page_range(range_str, max_pages)
                
                for p in pages_to_include:
                    merger.append(f, pages=(p, p+1))
                        
            elif ext == 'docx':
                try:
                    text = docx_to_text(f)
                    if text.strip():
                        docx_pdf_buf = text_to_pdf_buffer(text)
                        merger.append(docx_pdf_buf)
                except Exception as docx_err:
                    print(f"Error merging docx {f.name}: {docx_err}")
                    continue

        output = io.BytesIO()
        merger.write(output)
        return output.getvalue()
    finally:
        merger.close()

def merge_csvs(files):
    output = io.StringIO()
    writer = csv.writer(output)
    for i, f in enumerate(files):
        content = f.read().decode('utf-8')
        f.seek(0)
        reader = csv.reader(io.StringIO(content))
        header = next(reader, None)
        if i == 0 and header: writer.writerow(header)
        for row in reader: writer.writerow(row)
    return output.getvalue().encode('utf-8')

def merge_excels(files):
    new_wb = openpyxl.Workbook()
    default_sheet = new_wb.active
    new_wb.remove(default_sheet)
    for f in files:
        wb = openpyxl.load_workbook(f, data_only=True)
        for sheet_name in wb.sheetnames:
            source_sheet = wb[sheet_name]
            target_sheet_name = f"{f.name}_{sheet_name}"[:31]
            target_sheet = new_wb.create_sheet(title=target_sheet_name)
            for row in source_sheet.iter_rows(values_only=True):
                target_sheet.append(row)
    output = io.BytesIO()
    new_wb.save(output)
    return output.getvalue()

